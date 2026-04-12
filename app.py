from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
import sqlite3
import os
import json
import io
import urllib.request
import urllib.error
import time
from difflib import SequenceMatcher

app = Flask(__name__)

# Use /data on Render (persistent disk), local file otherwise
if os.path.isdir('/data'):
    DATABASE = '/data/golf_pool.db'
else:
    DATABASE = os.path.join(os.path.dirname(__file__), 'golf_pool.db')

ROUND_NAMES = ['Thursday', 'Friday', 'Saturday', 'Sunday']

# Cache ESPN responses to avoid hammering the API
_espn_cache = {'data': None, 'time': 0, 'event_id': None}
ESPN_CACHE_SECONDS = 60  # refresh every 60s


def get_db():
    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row
    return db


def init_db():
    db = get_db()
    db.executescript('''
        CREATE TABLE IF NOT EXISTS tournament (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            year INTEGER NOT NULL,
            espn_event_id TEXT
        );
        CREATE TABLE IF NOT EXISTS golfer (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL UNIQUE
        );
        CREATE TABLE IF NOT EXISTS pool_member (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            tournament_id INTEGER,
            FOREIGN KEY (tournament_id) REFERENCES tournament(id)
        );
        CREATE TABLE IF NOT EXISTS pick (
            id INTEGER PRIMARY KEY,
            pool_member_id INTEGER NOT NULL,
            golfer_id INTEGER NOT NULL,
            pick_order INTEGER NOT NULL,
            FOREIGN KEY (pool_member_id) REFERENCES pool_member(id),
            FOREIGN KEY (golfer_id) REFERENCES golfer(id),
            UNIQUE(pool_member_id, pick_order)
        );
        CREATE TABLE IF NOT EXISTS score (
            id INTEGER PRIMARY KEY,
            pick_id INTEGER NOT NULL,
            round_number INTEGER NOT NULL,
            score INTEGER,
            is_mc INTEGER DEFAULT 0,
            FOREIGN KEY (pick_id) REFERENCES pick(id),
            UNIQUE(pick_id, round_number)
        );
    ''')
    # Add espn_event_id column if missing (migration for existing DBs)
    try:
        db.execute('SELECT espn_event_id FROM tournament LIMIT 1')
    except sqlite3.OperationalError:
        db.execute('ALTER TABLE tournament ADD COLUMN espn_event_id TEXT')
    db.commit()
    db.close()


def seed_golfers():
    """Seed with a list of PGA Tour players."""
    golfers = [
        "Scottie Scheffler", "Xander Schauffele", "Rory McIlroy", "Collin Morikawa",
        "Ludvig Aberg", "Wyndham Clark", "Sahith Theegala", "Viktor Hovland",
        "Patrick Cantlay", "Hideki Matsuyama", "Tommy Fleetwood", "Shane Lowry",
        "Sungjae Im", "Tom Kim", "Russell Henley", "Sam Burns",
        "Keegan Bradley", "Justin Thomas", "Tony Finau", "Matt Fitzpatrick",
        "Robert MacIntyre", "Cameron Young", "Akshay Bhatia", "Corey Conners",
        "Sepp Straka", "Jason Day", "Jordan Spieth", "Adam Scott",
        "Max Homa", "Denny McCarthy", "Chris Kirk", "Brian Harman",
        "Cameron Smith", "Min Woo Lee", "Byeong Hun An", "Si Woo Kim",
        "Maverick McNealy", "Justin Rose", "Taylor Moore", "Davis Riley",
        "Rickie Fowler", "Billy Horschel", "JJ Spaun", "Austin Eckroat",
        "Christiaan Bezuidenhout", "Stephan Jaeger", "Jake Knapp", "Daniel Berger",
        "Brooks Koepka", "Bryson DeChambeau", "Dustin Johnson", "Phil Mickelson",
        "Jon Rahm", "Cameron Davis", "Nick Dunlap", "Ben Griffin",
        "Eric Cole", "Beau Hossler", "J.T. Poston", "Davis Thompson",
        "Nick Taylor", "Taylor Pendrith", "Luke Clanton", "Andrew Novak",
        "Kevin Yu", "Patrick Rodgers", "Harris English", "Joel Dahmen",
        "Nico Echavarria", "Austin Smotherman", "Chris Gotterup", "Jacob Bridgeman",
        "Mac Meissner", "Charley Hoffman", "Alex Noren", "Thomas Detry",
        "Emiliano Grillo", "Mark Hubbard", "Lee Hodges", "Mackenzie Hughes",
        "Keith Mitchell", "Doug Ghim", "S.H. Kim", "Matthieu Pavon",
        "Will Zalatoris", "Tyrrell Hatton", "Joaquin Niemann", "Abraham Ancer",
        "Adrian Meronk", "Mito Pereira", "C.T. Pan", "Harry Higgs",
        "Lucas Glover", "Kurt Kitayama", "Brendon Todd", "Kevin Kisner",
        "Patton Kizzire", "Webb Simpson", "Gary Woodland", "Zach Johnson",
    ]
    db = get_db()
    for name in golfers:
        db.execute('INSERT OR IGNORE INTO golfer (name) VALUES (?)', (name,))
    db.commit()
    db.close()


# ---------------------------------------------------------------------------
# ESPN Live Scores
# ---------------------------------------------------------------------------

def fetch_espn_scoreboard(espn_event_id=None):
    """Fetch live leaderboard data from ESPN. Returns parsed JSON or None.
    espn_event_id can be an event ID or a date string (YYYYMMDD).
    Dates work for past/completed tournaments; event IDs for current ones.
    """
    now = time.time()
    if (_espn_cache['data'] and
            now - _espn_cache['time'] < ESPN_CACHE_SECONDS and
            _espn_cache['event_id'] == espn_event_id):
        return _espn_cache['data']

    url = 'https://site.api.espn.com/apis/site/v2/sports/golf/pga/scoreboard'
    if espn_event_id:
        # If it looks like a date (8 digits), use dates= param; otherwise event=
        stripped = espn_event_id.strip()
        if stripped.isdigit() and len(stripped) == 8:
            url += f'?dates={stripped}'
        else:
            url += f'?event={stripped}'

    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'GolfPool/1.0'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode())
        _espn_cache['data'] = data
        _espn_cache['time'] = now
        _espn_cache['event_id'] = espn_event_id
        return data
    except (urllib.error.URLError, json.JSONDecodeError, OSError) as e:
        print(f"ESPN fetch error: {e}")
        return _espn_cache.get('data')  # return stale cache if available


def parse_espn_competitors(data):
    """Parse ESPN data into a dict of golfer_name -> {scores, status, position}."""
    if not data or 'events' not in data or not data['events']:
        return {}, None

    event = data['events'][0]
    tournament_status = event.get('status', {}).get('type', {}).get('state', 'pre')

    competitors = {}
    if 'competitions' not in event or not event['competitions']:
        return {}, tournament_status

    for comp in event['competitions'][0].get('competitors', []):
        athlete = comp.get('athlete', {})
        name = athlete.get('displayName', '')
        if not name:
            continue

        # Per-round scores from linescores
        # ESPN sends value=0 with displayValue="-" for unplayed rounds
        round_scores = {}
        missed_cut = False
        played_rounds = 0
        for ls in comp.get('linescores', []):
            rnd = ls.get('period', 0)
            val = ls.get('value')
            display = ls.get('displayValue', '')
            # Skip unplayed rounds (value=0, displayValue="-")
            if val is not None and val != 0:
                round_scores[rnd] = int(val)
                played_rounds += 1
            elif val == 0 and display not in ('-', '', None):
                # Genuine score of 0 (extremely unlikely in golf but handle it)
                round_scores[rnd] = 0
                played_rounds += 1

        # Detect missed cut
        status_detail = comp.get('status', {}).get('type', {}).get('description', '')
        if 'cut' in status_detail.lower():
            missed_cut = True
        # Also detect MC if player only has 2 real rounds and round 3 has started
        # (ESPN sometimes doesn't set the cut status text)
        elif played_rounds == 2 and len(comp.get('linescores', [])) >= 3:
            # Round 3 linescore exists but value is 0 with "-" display = cut
            ls3 = [ls for ls in comp.get('linescores', []) if ls.get('period') == 3]
            if ls3 and ls3[0].get('value') == 0 and ls3[0].get('displayValue', '') in ('-', '', None):
                missed_cut = True

        competitors[name] = {
            'round_scores': round_scores,
            'missed_cut': missed_cut,
            'total_score': comp.get('score', 'E'),
            'position': comp.get('sortOrder', 999),
            'status_detail': status_detail,
            'thru': comp.get('status', {}).get('displayValue', ''),
        }

    return competitors, tournament_status


def normalize_name(name):
    """Normalize a golfer name for comparison."""
    import unicodedata
    # Strip accents (Åberg -> Aberg)
    nfkd = unicodedata.normalize('NFKD', name)
    ascii_name = ''.join(c for c in nfkd if not unicodedata.combining(c))
    return ascii_name.lower().strip().replace('.', '').replace("'", '').replace('-', ' ')


def fuzzy_match_golfer(pick_name, espn_names, threshold=0.55):
    """Match a picked golfer name to an ESPN name using fuzzy matching.
    Handles abbreviations like 'S. Scheffler' -> 'Scottie Scheffler',
    'X. Schauffele' -> 'Xander Schauffele', 'M. Fitzpatrick' -> 'Matt Fitzpatrick'.
    """
    pick_norm = normalize_name(pick_name)

    # Try exact match first
    for espn_name in espn_names:
        if normalize_name(espn_name) == pick_norm:
            return espn_name

    pick_parts = pick_norm.split()
    if not pick_parts:
        return None

    pick_last = pick_parts[-1]

    # Handle abbreviated first names (e.g., "S Scheffler", "JJ Spaun", "M W Lee")
    # Get all initials from the pick name (everything except the last word)
    pick_initials = [p[0] for p in pick_parts[:-1] if p]
    has_abbreviation = all(len(p) <= 2 for p in pick_parts[:-1]) and len(pick_parts) >= 2

    if has_abbreviation and pick_initials:
        candidates = []
        for espn_name in espn_names:
            espn_norm = normalize_name(espn_name)
            espn_parts = espn_norm.split()
            if not espn_parts:
                continue
            espn_last = espn_parts[-1]

            # Last name must match
            if espn_last != pick_last:
                continue

            # Check if first initials match
            espn_initials = [p[0] for p in espn_parts[:-1] if p]
            if len(pick_initials) == len(espn_initials):
                if all(pi == ei for pi, ei in zip(pick_initials, espn_initials)):
                    candidates.append(espn_name)
            elif len(pick_initials) == 1 and espn_initials and pick_initials[0] == espn_initials[0]:
                # Single initial, first initial matches
                candidates.append(espn_name)

        if len(candidates) == 1:
            return candidates[0]
        if candidates:
            # Multiple matches with same initial+last - pick best fuzzy
            best = max(candidates, key=lambda c: SequenceMatcher(None, pick_norm, normalize_name(c)).ratio())
            return best

    # Last name only match (only 1 player with that last name)
    last_name_matches = [n for n in espn_names if normalize_name(n).split()[-1] == pick_last]
    if len(last_name_matches) == 1:
        return last_name_matches[0]

    # Fuzzy ratio match as fallback - but require last name to be close
    best_match = None
    best_ratio = 0
    for espn_name in espn_names:
        espn_norm = normalize_name(espn_name)
        espn_parts = espn_norm.split()
        espn_last = espn_parts[-1] if espn_parts else ''
        # Last names must be similar (prevents "berger" matching "aberg")
        last_ratio = SequenceMatcher(None, pick_last, espn_last).ratio()
        if last_ratio < 0.75:
            continue
        ratio = SequenceMatcher(None, pick_norm, espn_norm).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_match = espn_name
    if best_ratio >= threshold:
        return best_match

    return None


def build_live_leaderboard(tid):
    """Build leaderboard using live ESPN data merged with pool picks."""
    db = get_db()
    t = db.execute('SELECT * FROM tournament WHERE id=?', (tid,)).fetchone()
    if not t:
        db.close()
        return None, None, None

    espn_event_id = t['espn_event_id'] if 'espn_event_id' in t.keys() else None
    espn_data = fetch_espn_scoreboard(espn_event_id)
    competitors, tournament_status = parse_espn_competitors(espn_data)
    espn_names = list(competitors.keys())

    members = db.execute('SELECT * FROM pool_member WHERE tournament_id=? ORDER BY name', (tid,)).fetchall()

    leaderboard = []
    for member in members:
        picks = db.execute('''
            SELECT p.id, p.pick_order, g.name as golfer_name
            FROM pick p JOIN golfer g ON p.golfer_id = g.id
            WHERE p.pool_member_id = ?
            ORDER BY p.pick_order
        ''', (member['id'],)).fetchall()

        round_totals = []
        golfer_data = []

        for pick in picks:
            golfer_name = pick['golfer_name']
            espn_match = fuzzy_match_golfer(golfer_name, espn_names)

            scores = {}
            missed_cut = False
            espn_info = None

            if espn_match and espn_match in competitors:
                espn_info = competitors[espn_match]
                for rnd, score in espn_info['round_scores'].items():
                    scores[rnd] = {'score': score, 'is_mc': False}
                missed_cut = espn_info['missed_cut']
            else:
                # Fall back to DB scores
                db_scores = db.execute(
                    'SELECT round_number, score, is_mc FROM score WHERE pick_id=? ORDER BY round_number',
                    (pick['id'],)
                ).fetchall()
                for s in db_scores:
                    scores[s['round_number']] = {'score': s['score'], 'is_mc': bool(s['is_mc'])}
                    if s['is_mc']:
                        missed_cut = True

            golfer_data.append({
                'name': golfer_name,
                'espn_name': espn_match,
                'scores': scores,
                'missed_cut': missed_cut,
                'total_score': espn_info['total_score'] if espn_info else None,
                'thru': espn_info['thru'] if espn_info else None,
                'position': espn_info['position'] if espn_info else None,
            })

        # Calculate per-round averages (best 4 of 6)
        # Track which golfers' scores are used (for red highlighting)
        # For rounds 3-4, MC golfers are excluded. If fewer than 4 golfers
        # made the cut, the member is eliminated (round total = 'MC').
        member_eliminated = False
        for rnd in range(1, 5):
            scored = []  # list of (score, golfer_index)
            eligible_count = 0  # golfers who could play this round
            for gi, gd in enumerate(golfer_data):
                # Rounds 3-4: MC golfers can't play
                if gd['missed_cut'] and rnd > 2:
                    continue
                eligible_count += 1
                if rnd in gd['scores']:
                    s = gd['scores'][rnd]
                    if s.get('is_mc'):
                        continue
                    if s['score'] is not None:
                        scored.append((s['score'], gi))

            # For rounds 3-4, if fewer than 4 golfers made the cut, member is eliminated
            if rnd > 2 and eligible_count < 4:
                round_totals.append('MC')
                member_eliminated = True
            elif len(scored) >= 4:
                scored.sort(key=lambda x: x[0])
                best4 = scored[:4]
                round_totals.append(sum(s for s, _ in best4) / 4)
                for _, gi in best4:
                    golfer_data[gi].setdefault('used_rounds', set()).add(rnd)
            else:
                round_totals.append(None)

        final = None
        if member_eliminated:
            final = 'MC'
        else:
            valid_totals = [rt for rt in round_totals if rt is not None and rt != 'MC']
            if valid_totals:
                final = sum(valid_totals) / len(valid_totals)

        # Check if any golfer for this member is currently on the course
        any_live = any(
            g.get('thru') and g['thru'] not in ('F', '--', '')
            for g in golfer_data
        )

        leaderboard.append({
            'member': dict(member),
            'golfers': golfer_data,
            'round_totals': round_totals,
            'final': final,
            'any_live': any_live,
        })

    leaderboard.sort(key=lambda x: (x['final'] == 'MC', x['final'] is None, x['final'] if isinstance(x['final'], (int, float)) else 999))
    db.close()

    return leaderboard, tournament_status, t


# ---------------------------------------------------------------------------
# Seed data
# ---------------------------------------------------------------------------

def seed_pool_data():
    """Populate the database with The Masters 2026 pool data."""
    db = get_db()

    # Skip if already seeded
    if db.execute('SELECT COUNT(*) FROM pool_member').fetchone()[0] > 0:
        db.close()
        return False

    # Tournament
    db.execute('DELETE FROM tournament')
    db.execute("INSERT INTO tournament (id, name, year, espn_event_id) VALUES (1, 'The Masters', 2026, '401811941')")

    members = ['Griffin','GG','Ray','Debbie','Ryan O.','Josh B.','Elice','Jill','George','Manny',
               'Chris','Gary','Liz','Alton','Mike C.','Josh','Betty Anne','Jeff','Jake','Coach',
               'Zach','Richard','Rob','Kenny','Karen']

    picks = [
        ['C. Young','R. McIlroy','J. Rahm','S. Scheffler','J. Rahm','J. Rahm','S. Scheffler','X. Schauffele','S. Scheffler','R. McIlroy','L. Aberg','J. Rahm','C. Young','S. Scheffler','S. Scheffler','S. Scheffler','H. Matsuyama','S. Scheffler','B. Dechambeau','C. Young','S. Scheffler','L. Aberg','C. Young','S. Scheffler','S. Scheffler'],
        ['X. Schauffele','T. Fleetwood','L. Aberg','X. Schauffele','C. Young','C. Young','T. Fleetwood','L. Aberg','L. Aberg','L. Aberg','B. Dechambeau','X. Schauffele','T. Fleetwood','B. Dechambeau','X. Schauffele','X. Schauffele','X. Schauffele','X. Schauffele','X. Schauffele','X. Schauffele','J. Rahm','T. Fleetwood','J. Rahm','X. Schauffele','T. Fleetwood'],
        ['J. Day','M. W. Lee','J. Rose','J. Rose','M. W. Lee','M. Fitzpatrick','R. McIntyre','J. Day','A. Bhatia','M. W. Lee','A. Bhatia','C. Gutterup','C. Gutterup','M. Fitzpatrick','R. McIntyre','C. Gutterup','J. Day','C. Gutterup','S. Lowry','R. McIntyre','C. Gutterup','R. McIntyre','C. Gutterup','V. Hovland','R. McIntyre'],
        ['J. Rose','A. Bhatia','M. Fitzpatrick','M. Fitzpatrick','M. Fitzpatrick','A. Bhatia','S. W. Kim','J. Rose','R. McIntyre','J. Rose','J. Rose','S. W. Kim','R. McIntyre','J. Rose','J. Rose','J. Rose','M. Fitzpatrick','R. McIntyre','A. Bhatia','A. Bhatia','M. W. Lee','A. Bhatia','J. Rose','J. Rose','A. Bhatia'],
        ['A. Scott','A. Scott','B. Koepka','A. Scott','B. Koepka','J. Spaun','J. Spieth','J. Thomas','B. Koepka','McNealy','A. Scott','J. Spieth','J. Spieth','N. Hojgaard','J. Knapp','A. Scott','J. Spieth','S. Straka','S. Straka','P. Reed','A. Scott','R. Henley','B. Koepka','A. Scott','R. Henley'],
        ['M. Homa','P. Reed','P. Reed','J. Knapp','A. Scott','P. Reed','J. Spaun','M. Homa','J. Spaun','P. Reed','H. English','P. Reed','M. McNealy','C. Conners','P. Reed','C. Conners','J. Spaun','J. Spaun','C. Conners','C. Conners','P. Reed','C. Conners','P. Reed','S. Straka','P. Reed'],
    ]

    member_ids = {}
    for name in members:
        db.execute('INSERT INTO pool_member (name, tournament_id) VALUES (?, 1)', (name,))
        member_ids[name] = db.execute('SELECT last_insert_rowid()').fetchone()[0]

    for pick_order, pick_row in enumerate(picks):
        for i, golfer_name in enumerate(pick_row):
            row = db.execute('SELECT id FROM golfer WHERE name = ?', (golfer_name,)).fetchone()
            if row:
                golfer_id = row[0]
            else:
                db.execute('INSERT INTO golfer (name) VALUES (?)', (golfer_name,))
                golfer_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]
            db.execute('INSERT INTO pick (pool_member_id, golfer_id, pick_order) VALUES (?, ?, ?)',
                       (member_ids[members[i]], golfer_id, pick_order + 1))

    db.commit()
    db.close()
    return True


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    db = get_db()
    tournaments = db.execute('SELECT * FROM tournament ORDER BY year DESC').fetchall()
    db.close()
    return render_template('index.html', tournaments=tournaments)


@app.route('/tournament/new', methods=['GET', 'POST'])
def new_tournament():
    if request.method == 'POST':
        name = request.form['name']
        year = request.form['year']
        espn_event_id = request.form.get('espn_event_id', '').strip() or None
        db = get_db()
        db.execute('INSERT INTO tournament (name, year, espn_event_id) VALUES (?, ?, ?)',
                   (name, year, espn_event_id))
        db.commit()
        db.close()
        return redirect(url_for('index'))
    # Fetch upcoming ESPN events for the dropdown
    espn_events = []
    data = fetch_espn_scoreboard()
    if data and 'events' in data:
        for ev in data['events']:
            espn_events.append({'id': ev['id'], 'name': ev.get('name', 'Unknown')})
    if data and 'leagues' in data:
        for league in data['leagues']:
            for cal in league.get('calendar', []):
                espn_events.append({'id': cal.get('id', ''), 'name': cal.get('label', '')})
    return render_template('new_tournament.html', espn_events=espn_events)


@app.route('/tournament/<int:tid>')
def tournament(tid):
    leaderboard, tournament_status, t = build_live_leaderboard(tid)
    if not t:
        return redirect(url_for('index'))

    # Ranked version (sorted by final score, lowest first)
    leaderboard_ranked = sorted(leaderboard, key=lambda x: (x['final'] == 'MC', x['final'] is None, x['final'] if isinstance(x['final'], (int, float)) else 999))

    # Assign placements to each entry
    for idx, entry in enumerate(leaderboard_ranked):
        entry['placement'] = idx + 1 if entry['final'] is not None and entry['final'] != 'MC' else None

    # Find low round: lowest Sunday (round 4) average among members NOT in top 4
    low_round_winner = None
    low_round_value = None
    low_round_rnd = 4  # Sunday only
    for entry in leaderboard_ranked:
        if entry['placement'] is not None and entry['placement'] <= 4:
            continue  # skip top 4 finishers
        if len(entry['round_totals']) >= 4 and entry['round_totals'][3] is not None and entry['round_totals'][3] != 'MC':
            rt = entry['round_totals'][3]
            if low_round_value is None or rt < low_round_value:
                low_round_value = rt
                low_round_winner = entry['member']['name']

    # Tag the low round winner
    for entry in leaderboard:
        entry['low_round'] = (entry['member']['name'] == low_round_winner)
        entry['low_round_value'] = low_round_value if entry['low_round'] else None
        entry['low_round_rnd'] = low_round_rnd if entry['low_round'] else None

    # By-name version preserves original member order (alphabetical)
    leaderboard_by_name = sorted(leaderboard, key=lambda x: x['member']['name'])

    return render_template('tournament.html', tournament=t,
                           leaderboard=leaderboard,
                           leaderboard_ranked=leaderboard_ranked,
                           leaderboard_by_name=leaderboard_by_name,
                           low_round_winner=low_round_winner,
                           low_round_value=low_round_value,
                           low_round_rnd=low_round_rnd,
                           round_names=ROUND_NAMES, tournament_status=tournament_status)


@app.route('/tournament/<int:tid>/export')
def export_excel(tid):
    """Export tournament data as an Excel file matching the spreadsheet format."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    leaderboard, tournament_status, t = build_live_leaderboard(tid)
    if not t:
        return redirect(url_for('index'))

    # Sort alphabetically like the original sheet
    entries = sorted(leaderboard, key=lambda x: x['member']['name'])

    # Compute placements
    ranked = sorted(leaderboard, key=lambda x: (x['final'] == 'MC', x['final'] is None, x['final'] if isinstance(x['final'], (int, float)) else 999))
    placement_map = {}
    for idx, entry in enumerate(ranked):
        if entry['final'] is not None and entry['final'] != 'MC':
            placement_map[entry['member']['name']] = idx + 1

    # Low round (Sunday only, outside top 4)
    lr_winner = None
    lr_value = None
    for entry in ranked:
        p = placement_map.get(entry['member']['name'])
        if p is not None and p <= 4:
            continue
        if len(entry['round_totals']) >= 4 and entry['round_totals'][3] is not None and entry['round_totals'][3] != 'MC':
            rt = entry['round_totals'][3]
            if lr_value is None or rt < lr_value:
                lr_value = rt
                lr_winner = entry['member']['name']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{t['name']} {t['year']}"

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='2D5A3D')
    round_font = Font(bold=True, color='FFD700', size=11)
    round_fill = PatternFill('solid', fgColor='1A472A')
    total_fill = PatternFill('solid', fgColor='E8F5E9')
    total_font = Font(bold=True, size=10)
    final_fill = PatternFill('solid', fgColor='FFD700')
    final_font = Font(bold=True, size=11)
    used_font = Font(bold=True, color='DC3545')  # red for counted scores
    mc_font = Font(bold=True, color='DC3545')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
    place_fills = {
        1: PatternFill('solid', fgColor='FFD700'),
        2: PatternFill('solid', fgColor='E0E0E0'),
        3: PatternFill('solid', fgColor='F0D0A0'),
        4: PatternFill('solid', fgColor='D4EDDA'),
    }
    lr_fill = PatternFill('solid', fgColor='E8DAEF')

    # --- Header row (row 1): member names ---
    row = 1
    for mi, entry in enumerate(entries):
        name_col = 2 + mi * 2      # B, D, F, ...
        score_col = name_col + 1    # C, E, G, ...
        cell = ws.cell(row=row, column=name_col, value=entry['member']['name'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        cell2 = ws.cell(row=row, column=score_col, value='Score')
        cell2.font = header_font
        cell2.fill = header_fill
        cell2.alignment = Alignment(horizontal='center')
        cell2.border = thin_border

    row = 2

    # --- 4 round blocks ---
    for rnd_idx, rnd_name in enumerate(ROUND_NAMES):
        rnd = rnd_idx + 1

        # Round label row
        label_cell = ws.cell(row=row, column=1, value=rnd_name)
        label_cell.font = round_font
        label_cell.fill = round_fill
        label_cell.border = thin_border
        for mi, entry in enumerate(entries):
            nc = 2 + mi * 2
            sc = nc + 1
            ws.cell(row=row, column=nc).fill = round_fill
            ws.cell(row=row, column=nc).border = thin_border
            ws.cell(row=row, column=sc).fill = round_fill
            ws.cell(row=row, column=sc).border = thin_border
        row += 1

        # 6 golfer rows
        for pick_idx in range(6):
            for mi, entry in enumerate(entries):
                nc = 2 + mi * 2
                sc = nc + 1
                if pick_idx < len(entry['golfers']):
                    g = entry['golfers'][pick_idx]
                    s = g['scores'].get(rnd)
                    is_used = rnd in g.get('used_rounds', set())

                    name_cell = ws.cell(row=row, column=nc, value=g['name'])
                    name_cell.border = thin_border

                    if s:
                        if s.get('is_mc') or (g['missed_cut'] and rnd > 2):
                            score_cell = ws.cell(row=row, column=sc, value='MC')
                            score_cell.font = mc_font
                        elif s['score'] is not None:
                            score_cell = ws.cell(row=row, column=sc, value=s['score'])
                            if is_used:
                                score_cell.font = used_font
                        else:
                            score_cell = ws.cell(row=row, column=sc)
                    elif g['missed_cut'] and rnd > 2:
                        score_cell = ws.cell(row=row, column=sc, value='MC')
                        score_cell.font = mc_font
                    else:
                        score_cell = ws.cell(row=row, column=sc)

                    score_cell.alignment = Alignment(horizontal='center')
                    score_cell.border = thin_border
            row += 1

        # TOTAL row
        total_label = ws.cell(row=row, column=1, value='TOTAL')
        total_label.font = total_font
        total_label.fill = total_fill
        total_label.border = thin_border
        for mi, entry in enumerate(entries):
            nc = 2 + mi * 2
            sc = nc + 1
            ws.cell(row=row, column=nc).fill = total_fill
            ws.cell(row=row, column=nc).border = thin_border
            rt = entry['round_totals'][rnd_idx]
            tc = ws.cell(row=row, column=sc, value='MC' if rt == 'MC' else (round(rt, 2) if rt is not None else None))
            tc.font = total_font
            tc.fill = total_fill
            tc.alignment = Alignment(horizontal='center')
            tc.border = thin_border
        row += 1

    # --- Placement row ---
    ws.cell(row=row, column=1).border = thin_border
    for mi, entry in enumerate(entries):
        nc = 2 + mi * 2
        sc = nc + 1
        p = placement_map.get(entry['member']['name'])
        is_lr = entry['member']['name'] == lr_winner

        label = ''
        if p == 1:
            label = '1st'
        elif p == 2:
            label = '2nd'
        elif p == 3:
            label = '3rd'
        elif p == 4:
            label = '4th'
        if is_lr:
            label = (label + ' ' if label else '') + 'LR'

        cell = ws.cell(row=row, column=nc, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.merge_cells(start_row=row, start_column=nc, end_row=row, end_column=sc)

        fill = None
        if p in place_fills:
            fill = place_fills[p]
        elif is_lr:
            fill = lr_fill
        if fill:
            cell.fill = fill
            ws.cell(row=row, column=sc).fill = fill
        ws.cell(row=row, column=sc).border = thin_border
    row += 1

    # --- FINAL row ---
    final_label = ws.cell(row=row, column=1, value='FINAL')
    final_label.font = final_font
    final_label.fill = final_fill
    final_label.border = thin_border
    for mi, entry in enumerate(entries):
        nc = 2 + mi * 2
        sc = nc + 1
        ws.cell(row=row, column=nc).fill = final_fill
        ws.cell(row=row, column=nc).border = thin_border
        fc = ws.cell(row=row, column=sc,
                     value='MC' if entry['final'] == 'MC' else (round(entry['final'], 2) if entry['final'] is not None else None))
        fc.font = final_font
        fc.fill = final_fill
        fc.alignment = Alignment(horizontal='center')
        fc.border = thin_border
    row += 1

    # --- Footer: names again ---
    for mi, entry in enumerate(entries):
        nc = 2 + mi * 2
        sc = nc + 1
        cell = ws.cell(row=row, column=nc, value=entry['member']['name'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.cell(row=row, column=sc).fill = header_fill
        ws.cell(row=row, column=sc).border = thin_border

    # Auto-fit column widths
    ws.column_dimensions['A'].width = 12
    for mi in range(len(entries)):
        nc_letter = openpyxl.utils.get_column_letter(2 + mi * 2)
        sc_letter = openpyxl.utils.get_column_letter(3 + mi * 2)
        ws.column_dimensions[nc_letter].width = 16
        ws.column_dimensions[sc_letter].width = 8

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{t['name'].replace(' ', '_')}_{t['year']}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@app.route('/api/live/<int:tid>')
def api_live(tid):
    """JSON API for live leaderboard data (used by auto-refresh)."""
    leaderboard, tournament_status, t = build_live_leaderboard(tid)
    if not t:
        return jsonify({'error': 'not found'}), 404

    return jsonify({
        'tournament_status': tournament_status,
        'last_updated': time.strftime('%I:%M:%S %p'),
        'leaderboard': [{
            'rank': idx + 1,
            'name': entry['member']['name'],
            'golfers': [{
                'name': g['name'],
                'espn_name': g.get('espn_name'),
                'missed_cut': g['missed_cut'],
                'total_score': g.get('total_score'),
                'thru': g.get('thru'),
                'position': g.get('position'),
                'scores': {str(k): v for k, v in g['scores'].items()},
            } for g in entry['golfers']],
            'round_totals': entry['round_totals'],
            'final': entry['final'],
        } for idx, entry in enumerate(leaderboard)]
    })


@app.route('/tournament/<int:tid>/member/new', methods=['GET', 'POST'])
def new_member(tid):
    if request.method == 'POST':
        name = request.form['name']
        db = get_db()
        db.execute('INSERT INTO pool_member (name, tournament_id) VALUES (?, ?)', (name, tid))
        db.commit()
        db.close()
        return redirect(url_for('enter_picks', tid=tid, member_name=name))
    db = get_db()
    t = db.execute('SELECT * FROM tournament WHERE id=?', (tid,)).fetchone()
    db.close()
    return render_template('new_member.html', tournament=t)


@app.route('/tournament/<int:tid>/picks', methods=['GET', 'POST'])
def enter_picks(tid):
    db = get_db()
    t = db.execute('SELECT * FROM tournament WHERE id=?', (tid,)).fetchone()
    members = db.execute('SELECT * FROM pool_member WHERE tournament_id=? ORDER BY name', (tid,)).fetchall()
    golfers = db.execute('SELECT * FROM golfer ORDER BY name').fetchall()

    if request.method == 'POST':
        member_id = request.form['member_id']
        # Clear existing picks
        db.execute('DELETE FROM score WHERE pick_id IN (SELECT id FROM pick WHERE pool_member_id=?)', (member_id,))
        db.execute('DELETE FROM pick WHERE pool_member_id=?', (member_id,))
        for i in range(1, 7):
            golfer_id = request.form.get(f'golfer_{i}')
            if golfer_id:
                db.execute('INSERT INTO pick (pool_member_id, golfer_id, pick_order) VALUES (?, ?, ?)',
                           (member_id, golfer_id, i))
        db.commit()
        db.close()
        return redirect(url_for('tournament', tid=tid))

    selected_member = request.args.get('member_id')
    existing_picks = []
    if selected_member:
        existing_picks = db.execute('''
            SELECT p.pick_order, p.golfer_id FROM pick p
            WHERE p.pool_member_id=? ORDER BY p.pick_order
        ''', (selected_member,)).fetchall()

    db.close()
    return render_template('enter_picks.html', tournament=t, members=members,
                           golfers=golfers, selected_member=selected_member,
                           existing_picks={p['pick_order']: p['golfer_id'] for p in existing_picks})


@app.route('/tournament/<int:tid>/scores', methods=['GET', 'POST'])
def enter_scores(tid):
    db = get_db()
    t = db.execute('SELECT * FROM tournament WHERE id=?', (tid,)).fetchone()
    round_num = int(request.args.get('round', 1))

    if request.method == 'POST':
        round_num = int(request.form['round_number'])
        for key, value in request.form.items():
            if key.startswith('score_'):
                pick_id = int(key.split('_')[1])
                is_mc = request.form.get(f'mc_{pick_id}') == '1'
                score_val = None if is_mc or value == '' else int(value)
                db.execute('''
                    INSERT INTO score (pick_id, round_number, score, is_mc)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(pick_id, round_number) DO UPDATE SET score=?, is_mc=?
                ''', (pick_id, round_num, score_val, int(is_mc), score_val, int(is_mc)))
        db.commit()
        db.close()
        return redirect(url_for('tournament', tid=tid))

    members = db.execute('''
        SELECT pm.id as member_id, pm.name as member_name,
               p.id as pick_id, p.pick_order, g.name as golfer_name
        FROM pool_member pm
        JOIN pick p ON p.pool_member_id = pm.id
        JOIN golfer g ON p.golfer_id = g.id
        WHERE pm.tournament_id = ?
        ORDER BY pm.name, p.pick_order
    ''', (tid,)).fetchall()

    existing_scores = {}
    for m in members:
        s = db.execute('SELECT score, is_mc FROM score WHERE pick_id=? AND round_number=?',
                       (m['pick_id'], round_num)).fetchone()
        if s:
            existing_scores[m['pick_id']] = {'score': s['score'], 'is_mc': s['is_mc']}

    from collections import OrderedDict
    grouped = OrderedDict()
    for m in members:
        if m['member_name'] not in grouped:
            grouped[m['member_name']] = []
        grouped[m['member_name']].append(dict(m))

    db.close()
    return render_template('enter_scores.html', tournament=t, grouped=grouped,
                           round_num=round_num, round_names=ROUND_NAMES,
                           existing_scores=existing_scores)


@app.route('/tournament/<int:tid>/sync', methods=['POST'])
def sync_espn_scores(tid):
    """Pull ESPN scores and write them into the DB (one-time sync)."""
    db = get_db()
    t = db.execute('SELECT * FROM tournament WHERE id=?', (tid,)).fetchone()
    espn_event_id = t['espn_event_id'] if 'espn_event_id' in t.keys() else None
    espn_data = fetch_espn_scoreboard(espn_event_id)
    competitors, _ = parse_espn_competitors(espn_data)
    espn_names = list(competitors.keys())

    members = db.execute('SELECT * FROM pool_member WHERE tournament_id=?', (tid,)).fetchall()
    synced = 0
    for member in members:
        picks = db.execute('''
            SELECT p.id, g.name as golfer_name
            FROM pick p JOIN golfer g ON p.golfer_id = g.id
            WHERE p.pool_member_id = ?
        ''', (member['id'],)).fetchall()

        for pick in picks:
            espn_match = fuzzy_match_golfer(pick['golfer_name'], espn_names)
            if espn_match and espn_match in competitors:
                info = competitors[espn_match]
                for rnd, score in info['round_scores'].items():
                    is_mc = 1 if info['missed_cut'] and rnd == max(info['round_scores'].keys()) else 0
                    db.execute('''
                        INSERT INTO score (pick_id, round_number, score, is_mc)
                        VALUES (?, ?, ?, ?)
                        ON CONFLICT(pick_id, round_number) DO UPDATE SET score=?, is_mc=?
                    ''', (pick['id'], rnd, score, is_mc, score, is_mc))
                    synced += 1
                # Mark MC for rounds they didn't play
                if info['missed_cut']:
                    for rnd in range(3, 5):
                        if rnd not in info['round_scores']:
                            db.execute('''
                                INSERT INTO score (pick_id, round_number, score, is_mc)
                                VALUES (?, ?, NULL, 1)
                                ON CONFLICT(pick_id, round_number) DO UPDATE SET score=NULL, is_mc=1
                            ''', (pick['id'], rnd))
                            synced += 1

    db.commit()
    db.close()
    return redirect(url_for('tournament', tid=tid))


@app.route('/api/golfers')
def api_golfers():
    q = request.args.get('q', '')
    db = get_db()
    if q:
        golfers = db.execute('SELECT id, name FROM golfer WHERE name LIKE ? ORDER BY name LIMIT 20',
                             (f'%{q}%',)).fetchall()
    else:
        golfers = db.execute('SELECT id, name FROM golfer ORDER BY name').fetchall()
    db.close()
    return jsonify([{'id': g['id'], 'name': g['name']} for g in golfers])


@app.route('/api/espn/events')
def api_espn_events():
    """List available ESPN events (for linking tournaments)."""
    data = fetch_espn_scoreboard()
    events = []
    if data and 'events' in data:
        for ev in data['events']:
            events.append({
                'id': ev['id'],
                'name': ev.get('name', ''),
                'status': ev.get('status', {}).get('type', {}).get('state', ''),
            })
    return jsonify(events)


@app.route('/golfer/add', methods=['POST'])
def add_golfer():
    name = request.form.get('name', '').strip()
    if name:
        db = get_db()
        try:
            db.execute('INSERT INTO golfer (name) VALUES (?)', (name,))
            db.commit()
        except sqlite3.IntegrityError:
            pass
        db.close()
    return redirect(request.referrer or url_for('index'))


@app.route('/tournament/<int:tid>/member/<int:mid>/delete', methods=['POST'])
def delete_member(tid, mid):
    db = get_db()
    db.execute('DELETE FROM score WHERE pick_id IN (SELECT id FROM pick WHERE pool_member_id=?)', (mid,))
    db.execute('DELETE FROM pick WHERE pool_member_id=?', (mid,))
    db.execute('DELETE FROM pool_member WHERE id=?', (mid,))
    db.commit()
    db.close()
    return redirect(url_for('tournament', tid=tid))


init_db()
seed_golfers()
seed_pool_data()

if __name__ == '__main__':
    app.run(debug=True, port=5001)

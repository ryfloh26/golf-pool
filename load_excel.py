"""Load the Golf Pool 2026 Excel data into the database."""
import sqlite3
import openpyxl
import os

DATABASE = os.path.join(os.path.dirname(__file__), 'golf_pool.db')

def get_db():
    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row
    return db

def ensure_golfer(db, name):
    """Get or create a golfer by name, return id."""
    row = db.execute('SELECT id FROM golfer WHERE name = ?', (name,)).fetchone()
    if row:
        return row['id']
    db.execute('INSERT INTO golfer (name) VALUES (?)', (name,))
    db.commit()
    return db.execute('SELECT id FROM golfer WHERE name = ?', (name,)).fetchone()['id']

def load():
    wb = openpyxl.load_workbook('/Users/ryanohara/Downloads/Golf Pool 2026.xlsx')
    ws = wb['Sheet1']
    db = get_db()

    # Create tournament
    db.execute('DELETE FROM score')
    db.execute('DELETE FROM pick')
    db.execute('DELETE FROM pool_member')
    db.execute('DELETE FROM tournament')
    db.execute("INSERT INTO tournament (id, name, year) VALUES (1, 'The Players Championship', 2026)")
    db.commit()

    # Parse header row (row 1) to get member names and their column positions
    # Format: col B=name, C=Score, D=name, E=Score, etc.
    members = []
    row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]
    col = 1  # B = index 1
    while col < len(row1):
        cell = row1[col]
        if cell.value and cell.value != 'Score':
            member_name = str(cell.value).strip()
            members.append({'name': member_name, 'name_col': col, 'score_col': col + 1})
            col += 2
        else:
            col += 1

    print(f"Found {len(members)} pool members: {[m['name'] for m in members]}")

    # Create pool members
    for m in members:
        db.execute('INSERT INTO pool_member (name, tournament_id) VALUES (?, 1)', (m['name'],))
        m['id'] = db.execute('SELECT id FROM pool_member WHERE name=? AND tournament_id=1',
                             (m['name'],)).fetchone()['id']
    db.commit()

    # Parse rounds - each round block: day label row, 6 golfer rows, TOTAL row
    # Rows: 2-8=Thu(+TOTAL at 8), 9-15=Fri, 16-22=Sat, 23-29=Sun
    round_starts = [2, 9, 16, 23]  # 1-indexed row numbers for day labels
    round_names = ['Thursday', 'Friday', 'Saturday', 'Sunday']

    for round_idx, start_row in enumerate(round_starts):
        round_num = round_idx + 1
        print(f"\nRound {round_num} ({round_names[round_idx]}), starting at row {start_row}")

        # Rows start_row through start_row+5 are the 6 golfer picks
        for pick_order in range(6):
            row_num = start_row + pick_order
            row = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=False))[0]

            for m in members:
                golfer_cell = row[m['name_col']]
                score_cell = row[m['score_col']]

                golfer_name = golfer_cell.value
                score_val = score_cell.value

                if not golfer_name:
                    continue

                golfer_name = str(golfer_name).strip()

                # Get or create golfer
                golfer_id = ensure_golfer(db, golfer_name)

                # Create pick if first round
                if round_num == 1:
                    db.execute('INSERT OR IGNORE INTO pick (pool_member_id, golfer_id, pick_order) VALUES (?, ?, ?)',
                               (m['id'], golfer_id, pick_order + 1))
                    db.commit()

                # Find pick_id
                pick = db.execute(
                    'SELECT id FROM pick WHERE pool_member_id=? AND pick_order=?',
                    (m['id'], pick_order + 1)
                ).fetchone()

                if not pick:
                    continue

                # Insert score
                is_mc = 1 if score_val == 'MC' else 0
                score_int = None if is_mc or score_val is None else int(score_val)

                db.execute('''
                    INSERT OR REPLACE INTO score (pick_id, round_number, score, is_mc)
                    VALUES (?, ?, ?, ?)
                ''', (pick['id'], round_num, score_int, is_mc))

        db.commit()
        print(f"  Loaded scores for round {round_num}")

    db.close()
    print("\nDone! All data loaded.")

if __name__ == '__main__':
    load()
